import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

path = '/home/sebastian/workspace-cs/matrix-multiplication-benchmark/'
algorithm = "MM_Posix/"
program = 'MM1p'
N = 30
sizes = [ "500", "1000", "1200", "2000"] 
threads = ['1', '2', '4', '8']

def readData(data, thread, start):
    for i, size in enumerate(sizes, start=1):
        if i != 0:
            f = open(f"{path}{algorithm}outputs/{program}-size{size}-T{thread}.txt")
            for j, l in enumerate(f, start=N*start):
                data[j][i] = float(l)


def main():
    data = np.zeros([N * len(sizes), len(threads)+1])

    if os.path.exists(path+"parcial2.xlsx"):
        book = load_workbook(path+"parcial2.xlsx")
    else:
        book = Workbook()
      
    try:
        writer = pd.ExcelWriter(path+"parcial2.xlsx", engine = 'openpyxl')
        writer.book = book

        for i, thread in enumerate(threads):
            readData(data, thread, i)
            data[i*N:, 0] = thread
        
        data /= 1000000

        for i, thread in enumerate(threads):
            data[i*N:i*N+30, 0] = thread
        
        df = pd.DataFrame(data=data, columns=["Threads"]+sizes)
        means = df.groupby("Threads").mean()

        df.to_excel(writer, sheet_name=program+"-values")
        means.to_excel(writer, sheet_name=program+"results")

        writer.save()
        writer.close()

    except Exception as e:
        print(e)


if __name__ == '__main__':
    main()
